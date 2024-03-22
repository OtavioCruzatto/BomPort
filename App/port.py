#######################################################################################################################

print("Importando bibliotecas....................", end="")
from openpyxl import load_workbook
from classes import Component
from openpyxl.styles import Alignment
from mapping import CODE, DESCRIPTION, QUANTITY, DESIGNATOR, PART_NUMBER_1, MANUFACTURER_1, PART_NUMBER_2, \
    MANUFACTURER_2, PART_NUMBER_3, MANUFACTURER_3, PART_NUMBER_4, MANUFACTURER_4
print("OK")

#######################################################################################################################

filename_bom_template = "Template.xlsx"
filename_bom_input = "Input.xlsx"
filename_bom_portada = "Output.xlsx"
linha_editavel = 22 # As alterações no excel começarão a partir desta linha
contador_de_componentes = 2

#######################################################################################################################

print("Carregando BOM Template...................", end="")
workbook_bom_template = load_workbook(filename="..\\Template\\" + filename_bom_template)
worksheet_bom_template = workbook_bom_template.active
print("OK")

#######################################################################################################################

print("Carregando BOM de Entrada.....................", end="")
workbook_bom_input = load_workbook(filename="..\\Input\\" + filename_bom_input,
                                   read_only=True)
worksheet_bom_input = workbook_bom_input.active
print("OK")

#######################################################################################################################

print("Criando BOM de saida atraves do Template...", end="")
workbook_bom_template.save(filename="..\\Output\\" + filename_bom_portada)
print("OK")

#######################################################################################################################

print("Carregando BOM de saida....................", end="")
workbook_bom_portada = load_workbook(filename="..\\Output\\" + filename_bom_portada)
worksheet_bom_portada = workbook_bom_portada.active
print("OK")

#######################################################################################################################

print("Carregando dados da BOM de entrada............", end="")

components = []

for row in worksheet_bom_input.iter_rows(min_row=2, values_only=True):
    component = Component(code=row[CODE-1],
                          description=row[DESCRIPTION-1],
                          quantity=row[QUANTITY-1],
                          designator=row[DESIGNATOR-1],
                          part_number_1=row[PART_NUMBER_1-1],
                          manufacturer_1=row[MANUFACTURER_1-1],
                          part_number_2=row[PART_NUMBER_2-1],
                          manufacturer_2=row[MANUFACTURER_2-1],
                          part_number_3=row[PART_NUMBER_3-1],
                          manufacturer_3=row[MANUFACTURER_3-1],
                          part_number_4=row[PART_NUMBER_4-1],
                          manufacturer_4=row[MANUFACTURER_4-1])
    components.append(component)
print("OK")

#######################################################################################################################

for component in components:

    print("Adicionando componente " + str(contador_de_componentes) + "..................", end="")
    worksheet_bom_portada.insert_rows(idx=linha_editavel, amount=4)

    worksheet_bom_portada.cell(row=linha_editavel, column=2, value=str(contador_de_componentes))
    worksheet_bom_portada.cell(row=linha_editavel, column=3, value=str(component.code))
    worksheet_bom_portada.cell(row=linha_editavel, column=4, value=str(component.description))
    worksheet_bom_portada.cell(row=linha_editavel, column=5, value=str(component.quantity))
    worksheet_bom_portada.cell(row=linha_editavel, column=6, value=str(component.designator))
    worksheet_bom_portada.cell(row=linha_editavel, column=7, value=str(component.part_number_1))
    worksheet_bom_portada.cell(row=linha_editavel, column=8, value=str(component.manufacturer_1))
    worksheet_bom_portada.cell(row=linha_editavel+1, column=7, value=str(component.part_number_2))
    worksheet_bom_portada.cell(row=linha_editavel+1, column=8, value=str(component.manufacturer_2))
    worksheet_bom_portada.cell(row=linha_editavel+2, column=7, value=str(component.part_number_3))
    worksheet_bom_portada.cell(row=linha_editavel+2, column=8, value=str(component.manufacturer_3))
    worksheet_bom_portada.cell(row=linha_editavel+3, column=7, value=str(component.part_number_4))
    worksheet_bom_portada.cell(row=linha_editavel+3, column=8, value=str(component.manufacturer_4))
    worksheet_bom_portada.cell(row=linha_editavel, column=9, value=str("N/A"))
    print("OK")

    print("Centralizando textos do componente " + str(contador_de_componentes) + "......", end="")
    worksheet_bom_portada.cell(row=linha_editavel, column=2).alignment = Alignment(horizontal="center", vertical="center")
    worksheet_bom_portada.cell(row=linha_editavel, column=3).alignment = Alignment(horizontal="center", vertical="center")
    worksheet_bom_portada.cell(row=linha_editavel, column=4).alignment = Alignment(vertical="center")
    worksheet_bom_portada.cell(row=linha_editavel, column=5).alignment = Alignment(horizontal="center", vertical="center")
    worksheet_bom_portada.cell(row=linha_editavel, column=6).alignment = Alignment(vertical="center")
    worksheet_bom_portada.cell(row=linha_editavel, column=7).alignment = Alignment(vertical="center")
    worksheet_bom_portada.cell(row=linha_editavel, column=8).alignment = Alignment(vertical="center")
    worksheet_bom_portada.cell(row=linha_editavel + 1, column=7).alignment = Alignment(vertical="center")
    worksheet_bom_portada.cell(row=linha_editavel + 1, column=8).alignment = Alignment(vertical="center")
    worksheet_bom_portada.cell(row=linha_editavel + 2, column=7).alignment = Alignment(vertical="center")
    worksheet_bom_portada.cell(row=linha_editavel + 2, column=8).alignment = Alignment(vertical="center")
    worksheet_bom_portada.cell(row=linha_editavel + 3, column=7).alignment = Alignment(vertical="center")
    worksheet_bom_portada.cell(row=linha_editavel + 3, column=8).alignment = Alignment(vertical="center")
    worksheet_bom_portada.cell(row=linha_editavel, column=9).alignment = Alignment(horizontal="center", vertical="center")
    print("OK")

    print("Mesclando celulas do componente " + str(contador_de_componentes) + ".........", end="")
    worksheet_bom_portada.merge_cells(start_row=linha_editavel, end_row=linha_editavel+3, start_column=2, end_column=2)
    worksheet_bom_portada.merge_cells(start_row=linha_editavel, end_row=linha_editavel+3, start_column=3, end_column=3)
    worksheet_bom_portada.merge_cells(start_row=linha_editavel, end_row=linha_editavel+3, start_column=4, end_column=4)
    worksheet_bom_portada.merge_cells(start_row=linha_editavel, end_row=linha_editavel+3, start_column=5, end_column=5)
    worksheet_bom_portada.merge_cells(start_row=linha_editavel, end_row=linha_editavel+3, start_column=6, end_column=6)
    worksheet_bom_portada.merge_cells(start_row=linha_editavel, end_row=linha_editavel+3, start_column=9, end_column=9)
    print("OK")

    linha_editavel = linha_editavel + 4
    contador_de_componentes = contador_de_componentes + 1

#######################################################################################################################

print("Salvando lista de saida.....................", end="")
workbook_bom_portada.save(filename="..\\Output\\" + filename_bom_portada)
print("OK")

#######################################################################################################################

input("Pressione enter para sair...")
print()
